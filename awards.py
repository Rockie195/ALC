from openpyxl import load_workbook
import os.path  #start file
import pandas as pd
import numpy as np
import sys

if len(sys.argv) != 4:
    print('Must provide file directory, date(mm/dd/yyyy), and program abbreviation')
    sys.exit()
try:
    df = pd.read_excel(sys.argv[1])
except FileNotFoundError:
    print("Provide an existing file path")
    sys.exit()

date = sys.argv[2]

program = sys.argv[3]

# Program abbreviations
if program == 'a':
    program = 'Academic Intensive English Program 10-weeks'
elif program == 'as':
    program = 'Academic Intensive English Program 12-weeks'
elif program == 'a6':
    program = 'Academic Intensive English Program 6-weeks'
elif program == 'i12':
    program = 'Intensive English Communication Program 12-weeks'
elif program == 's11':
    program = 'Intensive English Communication Program 11-weeks'
elif program == 'ia':
    program = 'Intensive English Communication Program - A 4-weeks'
elif program == 'ib':
    program = 'Intensive English Communication Program - B 4-weeks'
elif program == 'ic':
    program = 'Intensive English Communication Program - C 4-weeks'
elif program == 'sc':
    program = 'Intensive English Communication Program - C 3-weeks'
else:
    print('Please look up abbreviations')
    sys.exit()

# Select the columns I want to keep
df = df[['First Name', 'Middle Name', 'Last Name', 'Suffix', 'Student Number','Birth Date',
         'Preferred Email', 'Preferred Phone Number','Preferred Street 1', 'Preferred Street 2',
         'Preferred City', 'Preferred State/Province', 'Preferred Foreign State/Province',
         'Preferred Country', 'Preferred Zip/Postal Code']]

# Change Birth Date format to mm/dd/yyyy
df['Birth Date'] = pd.to_datetime(df['Birth Date'], errors='coerce')
df['Birth Date'] = df['Birth Date'].dt.strftime('%m/%d/%Y')

# Combine two columns and delete one
df['Preferred State/Province'] = df['Preferred State/Province'].fillna(df['Preferred Foreign State/Province'])
df.drop(['Preferred Foreign State/Province'], axis = 1, inplace = True)

# Replace null values with an asterisk
df['Preferred State/Province'].fillna('*', inplace = True)
df['Preferred Zip/Postal Code'].fillna('*', inplace = True)

# Add empty columns
df.insert(8, 'Hold', np.nan)
df['Expedited']=np.nan

# Add extra columns needed for upload
df['Earner Display Name'] = df['First Name'] + ' ' +df['Last Name']
df['Award Title'] = 'Award of Completion'
df['Award Date'] = date
df['Major'] = program
df['Honors'] = np.nan
df['Concentration'] = np.nan

# Rename columns to match excel sheet upload
df.rename(columns={'Suffix': 'Name Suffix', 'Student Number': 'Unique ID', 'Birth Date': 'Date of Birth',
                   'Preferred Email': 'Email', 'Preferred Phone Number': 'SMS Number',
                   'Preferred Street 1' : 'Address 1', 'Preferred Street 2': 'Address 2', 'Preferred City': 'City',
                   'Preferred State/Province': 'State / Province', 'Preferred Country': 'Country',
                   'Preferred Zip/Postal Code': 'ZIP / Postal Code'}, inplace = True)

# Write any changes to the excel sheet, save and open
book = load_workbook(sys.argv[1])
del book[book.sheetnames[0]]
writer = pd.ExcelWriter(sys.argv[1], engine = 'openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df.to_excel(writer, 'Sheet1', index = False)
writer.save()
os.startfile(sys.argv[1])
