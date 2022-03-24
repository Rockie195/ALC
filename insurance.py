from openpyxl import load_workbook
import os.path  #start file
import pandas as pd
import numpy as np
import sys

if len(sys.argv) != 4:
    print('Must provide file directory, start and end date(mm/dd/yyyy)')
    sys.exit()
try:
    df = pd.read_excel(sys.argv[1])
except FileNotFoundError:
    print("Provide an existing file path")
    sys.exit()

start_date = sys.argv[2]

end_date = sys.argv[3]

# Select the columns I want to keep
df = df[['Student Number', 'Last Name', 'First Name', 'Age', 'Birth Date', 'Gender',
         'Country of Citizenship','Preferred Email', 'Local Street 1', 'Local Street 2',
         'Local City', 'Local State/Province', 'Local Country', 'Local Zip/Postal Code']]

# Replace null values with an asterisk
df['Local Street 1'].fillna('10920 Lindbrook Dr', inplace = True)
df['Local City'].fillna('Los Angeles', inplace = True)
df['Local State/Province'].fillna('CA', inplace = True)
df['Local Country'].fillna('USA', inplace = True)
df['Local Zip/Postal Code'].fillna('90024', inplace = True)
# df['A'] = df['Local Street 1']
# Replace null values in Local Street 2 only if Local Street 1's location is Lindbrook
pairing = {'10920 Lindbrook Dr': '#100'}
mapping = df['Local Street 1'].map(pairing)
df['Local Street 2'] = df['Local Street 2'].combine_first(mapping)

# Add extra columns needed for upload
df['Start Date'] = start_date
df['End Date'] = end_date
print(df)
# Write any changes to the excel sheet, save and open
book = load_workbook(sys.argv[1])
del book[book.sheetnames[0]]
writer = pd.ExcelWriter(sys.argv[1], engine = 'openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
df.to_excel(writer, 'Sheet1', index = False)
writer.save()
os.startfile(sys.argv[1])
