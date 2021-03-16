import openpyxl
from tkinter import *
from tkinter import ttk
import os.path  #start file
import pandas as pd #see if pandas can read openpyxl workbook to make dataframes for IECP
import numpy as np
#from IPython.display import display
from pandastable import Table
import tabloo #display pandas dataframe if pandas reads openpyxl

# Check to see if the IECP workbook has been formatted
CLEANED_IECP_WORKBOOK = False

def pullist_IECP():
    file_location = e1.get()
    # Clean the workbook if it has not been cleaned
    if not CLEANED_IECP_WORKBOOK:
        clean_iecp_sheet(file_location)

    # Get the ids of students with bad grades
    sort_sheet = pd.read_excel(file_location, sheet_name='SORT')
    bad_grade = sort_sheet[np.in1d(sort_sheet['GRADE'], ['D', 'D+', 'D-', 'F', 'F*', 'NP', 'W', 'W*'])]
    unique_bad_grade_ids = pd.DataFrame({'ID': list(set(bad_grade['STUDENT ID']))})

    # Check if they failed over the allowed number of times to fail
    number_of_fails = int(tkvar.get())
    bad_grade['Fail'] = bad_grade['GRADE'] == "NP"
    many_fails = bad_grade.groupby(['STUDENT ID'])['Fail'].sum()[bad_grade.groupby(['STUDENT ID'])["Fail"].sum()>=number_of_fails]
    unique_bad_grade_ids = unique_bad_grade_ids[np.in1d(unique_bad_grade_ids['ID'], many_fails.index)]

    # Create the PULL sheet
    pull_sheet = sort_sheet[np.in1d(sort_sheet['STUDENT ID'],unique_bad_grade_ids['ID'])]
    iecp_pull_wb = openpyxl.load_workbook(file_location)
    writer = pd.ExcelWriter(file_location, engine='openpyxl')
    writer.book = iecp_pull_wb
    writer.sheets = dict((ws.title, ws) for ws in iecp_pull_wb.worksheets)
    pull_sheet.to_excel(writer, "PULL", index=False)

    # Create the PULL_SIMPLE sheet


    #print(bad_grade)
    #print(unique_bad_grade_ids)

    iecp_pull_wb.save(file_location)
    os.startfile(file_location)
    print("Working on it")
    return

# Display a data frame table to the user
def grd():
    file_location=e1.get()
    # Clean the workbook if it has not been cleaned
    if not CLEANED_IECP_WORKBOOK:
        clean_iecp_sheet(file_location)

    # From the SORT sheet, select the columns the user will need to see if a student received a letter grade
    sort_sheet = pd.read_excel(file_location, sheet_name='SORT')
    letter_grade_sheet = pd.read_excel(file_location, sheet_name='lettergrade')
    # Select the students that have the same id as one of the ids in the lettergrade sheet
    letter_grade = sort_sheet[np.in1d(sort_sheet['STUDENT ID'],letter_grade_sheet['X Number'])]
    letter_grade_columns = letter_grade[["STUDENT ID", "LAST NAME", "FIRST NAME", "PROJECT ID", "COURSE TITLE", "INSTRUCTOR", "GRADE"]]
    df = pd.DataFrame(letter_grade_columns)

    # Create a new TK GUI
    # Might be able to use toplevel(?) method, but this suffices
    new_root = Tk()
    new_root.geometry("")
    frame = ttk.Frame(new_root)
    frame.pack(fill='both', expand=True)
    pt = Table(frame, dataframe=df)
    pt.show()
    new_root.mainloop()
    return

def clean_iecp_sheet(file_location):
    iecp_pull_wb = openpyxl.load_workbook(file_location)

    # Format the first sheet
    ws_raw = iecp_pull_wb['Sheet - 1']
    ws_raw.title = 'RAW'
    ws_raw.merged_cells = []
    ws_raw.delete_rows(1, 9)

    # Duplicate the first sheet into the second index
    ws_sort = iecp_pull_wb.copy_worksheet(ws_raw)
    ws_sort.title = 'SORT'
    iecp_pull_wb.move_sheet("SORT", -(len(iecp_pull_wb.sheetnames) - 2))

    # Modify the letter grade sheet
    ws_lettergrade = iecp_pull_wb['Sheet1']
    ws_lettergrade.title = 'lettergrade'
    ws_lettergrade.insert_rows(1)
    ws_lettergrade.cell(row = 1, column = 1).value = 'X Number'

    # Save the cleaned excel document
    iecp_pull_wb.save(file_location)
    # Global variables are looked down on in python :)
    # User might try cleaning workbook more than once
    global CLEANED_IECP_WORKBOOK
    CLEANED_IECP_WORKBOOK = True
    return



######
#AIEP#
######
def pullist_AIEP():
    file_location = e2.get()
    aiep_pull_wb = openpyxl.load_workbook(file_location)

    #Format the first sheet
    ws_sort = aiep_pull_wb['Sheet - 1']
    ws_sort.title = 'SORT'
    ws_sort.merged_cells = []
    ws_sort.delete_rows(1, 9)

    #Create the attendance sheet
    ws_attendance = aiep_pull_wb['Sheet1']
    ws_attendance.merged_cells = []
    ws_attendance.delete_rows(1, 7)
    ws_attendance.delete_rows(2, 2)
    ws_attendance.delete_rows(len(ws_attendance['A']) - 3, 4)
    for col in ws_attendance.iter_cols(min_col=7, min_row=1, max_col=19):
        for cell in col:
            if cell.value == '-':
                cell.value = '0'
    #Create a dictionary that contains different attendance hours per student id
    student_cell = {}
    for row in ws_attendance.iter_rows():
        if row[0].value is not None:
            id_and_name_cleanup(str(row[0].value), row[12].value, row[14].value, float(row[12].value) + float(row[14].value), row[16].value, student_cell)
    #Create a header row and ATTENDANCE sheet
    first_row_headers = ['StudentID', 'LastName', 'FirstName', 'Absent Hours', 'Late Hours', 'Total Absent Hours', 'Excused']

    aiep_pull_wb.create_sheet('ATTENDANCE')
    ws_final_attendance = aiep_pull_wb['ATTENDANCE']
    for i, headers in enumerate(first_row_headers):
        ws_final_attendance.cell(row = 1, column = i + 1).value = headers
    #Populate the attendance sheet with the values inputted in the dictionary
    next_row = 2
    for key, values in student_cell.items():
        ws_final_attendance.cell(column = 1, row = next_row).value = key
        for i in range(2, 8):
            ws_final_attendance.cell(column = i, row = next_row).value = values[i - 2]
        next_row += 1
    #remove the original sheet with messy data
    aiep_pull_wb.remove(aiep_pull_wb["Sheet1"])

    #Create a unique set of student
    unique_ids = set()
    for row in ws_sort.iter_rows():
        if row[12].value != "STUDENT ID" and row[12].value is not None:
            unique_ids.add(row[12].value)

    no_certificate = {}
    for id in unique_ids:
        no_certificate[id] = False
    #Check to see if the student received one or more 'D','D+','D-', 'F', 'F*', 'W', 'W*'
    #Include them in the no certificate list
    fail = ['D','D+','D-', 'F', 'F*', 'W', 'W*']
    #ws_sort = aiep_pull_wb['SORT']
    for row in ws_sort.iter_rows():
        if row[18].value in fail:
            no_certificate[row[12].value] = True

    #Check to see if the student received 2 or more 'NP's
    many_nps = {}
    for id in unique_ids:
        many_nps[id] = 0
    for row in ws_sort.iter_rows():
        if row[18].value == "NP":
            many_nps[row[12].value] += 1

    for id in many_nps:
        if many_nps[id] > 1:
            no_certificate[id] = True

    #Get all of the information for students not receiving
    #a certificate due to grades from the SORT Sheet to create the PULL sheet
    pull_sheet_headers = []
    for col in ws_sort.iter_cols():
        if col[0].value is not None:
            pull_sheet_headers.append(col[0].value)

    aiep_pull_wb.create_sheet('PULL')
    ws_pull = aiep_pull_wb["PULL"]
    for i, headers in enumerate(pull_sheet_headers):
        ws_pull.cell(row = 1, column = i + 1).value = pull_sheet_headers[i]

    for row in ws_sort.iter_rows():
        if row[12].value != "STUDENT ID" and row[12].value is not None:
            if no_certificate[row[12].value]:
                ws_pull.append((cell.value for cell in row[:]))

    #Create the PULL_SIMPLE sheet
    pull_simple_headers = ['StudentID', 'LastName', 'FirstName', 'ProjectID', 'CourseTitle' ,'Instructor']
    aiep_pull_wb.create_sheet('PULL_SIMPLE')
    ws_pull_simple = aiep_pull_wb["PULL_SIMPLE"]
    for i, headers in enumerate(pull_simple_headers):
        ws_pull_simple.cell(row = 1, column = i + 1).value = pull_simple_headers[i]
    #Only list them once with the first class that appears for the sort sheet
    already_seen = {}
    for id in unique_ids:
        already_seen[id] = False

    for row in ws_sort.iter_rows():
        if row[12].value != "STUDENT ID" and row[12].value is not None and not already_seen[row[12].value]:
            already_seen[row[12].value] = True
            if no_certificate[row[12].value]:
                ws_pull_simple.append((cell.value for cell in [row[12], row[10],
                                                        row[11], row[2], row[4], row[9]]))

    #Check to see if they went over the required attendance limit
    failed_attendance = {}
    for id in unique_ids:
        failed_attendance[id] = False
    required_attendance = tkvar2.get()
    for row in ws_final_attendance.iter_rows():
        if row[5].value != "Total Absent Hours":
            if float(row[5].value) >= required_attendance:
                no_certificate[row[0].value] = True
                failed_attendance[row[0].value] = [row[1].value, row[2].value]

    #List those with attendance underneath the top ones
    for row in ws_final_attendance.iter_rows():
        if row[0].value != "StudentID" and row[0].value is not None:
            if failed_attendance[row[0].value]:
                ws_pull_simple.append((cell.value for cell in [row[0], row[1], row[2]]))


    #Create the PULL_PRINT sheet
    aiep_pull_wb.create_sheet('PULL_PRINT')
    ws_pull_print = aiep_pull_wb["PULL_PRINT"]

    for row in ws_pull.iter_rows():
        ws_pull_print.append((cell.value for cell in [row[2], row[4], row[9], row[10], row[11],
                                                      row[12], row[17], row[18]]))

    #Create the CHANGE sheet
    aiep_pull_wb.create_sheet('CHANGE')
    ws_change = aiep_pull_wb["CHANGE"]
    change_headers = ['COURSE NO.', 'SECTION NO.', 'PROJECT ID', 'ENROLLMENT STATUS', 'COURSE TITLE',
                      'START DATE', 'END DATE', 'BUILDING', 'ROOM', 'INSTRUCTOR', 'LAST NAME',
                      'FIRST NAME', 'STUDENT ID', 'EMAIL', 'GENDER', 'BIRTH DATE', 'CITIZENSHIP',
                      'ABSENCES', 'GRADE From', 'GRADE To']
    for i, headers in enumerate(change_headers):
        ws_change.cell(row = 1, column = i + 1).value = headers


    #Create the PROBATION sheet
    aiep_pull_wb.create_sheet('PROBATION')
    ws_probation = aiep_pull_wb['PROBATION']
    probation_headers = ['STUDENT ID', 'LAST NAME', 'FIRST NAME', 'Reason 1', 'Reason 2', 'DID NOT PASS',
                         'DID NOT PASS 2', 'DID NOT PASS 3', 'DID NOT PASS 4', 'Enrolled in SU?', 'COMMENTS']
    for i, headers in enumerate(probation_headers):
        ws_probation.cell(row = 1, column = i + 1).value = headers

    unique_pulls = set()
    seen_pulls = {}
    for row in ws_pull_simple.iter_rows():
        if row[0].value != "StudentID" and row[0].value is not None:
            unique_pulls.add(row[0].value)

    for id in unique_pulls:
        seen_pulls[id] = False

    for row in ws_pull_simple.iter_rows():
        if row[0].value != "StudentID":
            if not seen_pulls[row[0].value]:
                seen_pulls[row[0].value] = True
                ws_probation.append((cell.value for cell in [row[0], row[1], row[2]]))


    #Find all the students who have only had A's
    #Create the ALL A sheet and the VALEDICTORIAN sheet
    not_an_A = ['A-', 'B+', 'B', 'B-', 'C+', 'C', 'C-', 'D+', 'D', 'D-', 'F', 'NP']
    all_a_students = {}
    for id in unique_ids:
        all_a_students[id] = True

    for row in ws_sort.iter_rows():
        if row[18].value in not_an_A:
            all_a_students[row[12].value] = False

    aiep_pull_wb.create_sheet("All A")
    ws_all_a = aiep_pull_wb['All A']

    all_a_headers = []
    for col in ws_sort.iter_cols():
        if col[0].value is not None:
            all_a_headers.append(col[0].value)

    for i, headers in enumerate(all_a_headers):
        ws_all_a.cell(row = 1, column = i + 1).value = all_a_headers[i]

    a_only_grades = ['A', 'A+', 'P']
    for row in ws_sort.iter_rows():
        if row[12].value != "STUDENT ID" and row[12].value is not None:
            if all_a_students[row[12].value] and row[18].value in a_only_grades:
                ws_all_a.append((cell.value for cell in row[:]))

    #Use the a_students to form the VALEDICTORIAN sheet
    a_student_seen = {}
    for a_student in all_a_students.keys():
        if all_a_students[a_student]:
            a_student_seen[a_student] = False

    aiep_pull_wb.create_sheet("VALEDICTORIAN")
    ws_valedictorian = aiep_pull_wb['VALEDICTORIAN']

    valedictorian_headers = ['LAST NAME', 'FIRST NAME', 'STUDENT ID', 'EMAIL', 'GENDER', 'CITIZENSHIP',
                            'LEVEL', 'StudentID', 'Total Absent Hours']
    for i, headers in enumerate(valedictorian_headers):
        ws_valedictorian.cell(row = 1, column = i + 1).value = valedictorian_headers[i]

    for row in ws_all_a.iter_rows():
        if row[12].value != "STUDENT ID" and row[12].value is not None:
            if not a_student_seen[row[12].value]:
                a_student_seen[row[12].value] = True
                ws_valedictorian.append((cell.value for cell in [row[10], row[11], row[12], row[13],
                                                                 row[14], row[16], row[19], row[12]]))
                ws_valedictorian.cell(row = len(ws_valedictorian['A']), column = 9).value = student_cell[row[12].value][4]

    aiep_pull_wb.save(file_location)
    os.startfile(file_location)

    # Add core level sheets

    return



#Given different attendance values create a dictionary where the key is
#the student id and the value is an array with values unique to that student
def id_and_name_cleanup(student_cell, absent_hours_cell, late_hours_cell, total_absent_hours_cell, excused_cell, all_info):
    student_id = student_cell[student_cell.find("(") + 1:student_cell.find(")")]
    last_name = student_cell[0:student_cell.find(",")]
    first_name = student_cell[student_cell.find(",") + 2:student_cell.find("(") - 1]
    all_info[student_id] = [last_name, first_name, absent_hours_cell, late_hours_cell, total_absent_hours_cell, excused_cell]






#####
# GUI#
#####
root = Tk()
root.title("Pull List GUI")
root.resizable(False, False)

# Defines and places the notebook widget
nb = ttk.Notebook(root)
nb.grid(row=0, column=0, columnspan=50, rowspan=49, sticky='NESW')

# Adds tab 1 of the notebook
page1 = ttk.Frame(nb)
nb.add(page1, text='IECP')

# Adds tab 2 of the notebook
page2 = ttk.Frame(nb)
nb.add(page2, text='AIEP')

######################
# add widgets for IECP#
######################
Label(page1, text="File Location", justify=LEFT).grid(row=1, columnspan=1, sticky='w')
Label(page1, text="Number of NP", justify=LEFT).grid(row=2, columnspan=1, sticky='w')

e1 = Entry(page1)
e1.grid(row=1, column=1, sticky=W)
tkvar = StringVar(page1)
choices = {1, 2, 3, 4, 5}
tkvar.set(2)
OptionMenu(page1, tkvar, *choices).grid(row=2, column=1, sticky=W)

Button(page1, text='Pull', command=pullist_IECP).grid(row=3, column=1, sticky=W, pady=4)
Button(page1, text='GradeGroup Check', command=grd).grid(row=3, column=2, sticky=W, pady=4)

page1.grid_rowconfigure(0, weight=1)
page1.grid_columnconfigure(0, weight=1)

footer_label1 = Label(page1, text="by Christian", background="white")
footer_label1.grid(column=4, row=4)

######################
# add widgets for AIEP#
######################
Label(page2, text="File Location").grid(row=1, columnspan=1, sticky='w')
Label(page2, text="Max Absent Hours").grid(row=2, columnspan=1, sticky='w')
e2 = Entry(page2)
e2.grid(row=1, column=1, columnspan=1, sticky="w")

tkvar2 = DoubleVar(page2)
choices2 = {22, 40, 41, 47}
tkvar2.set(40)
OptionMenu(page2, tkvar2, *choices2).grid(row=2, column=1, columnspan=1, sticky="w")

Button(page2, text='Pull', command=pullist_AIEP).grid(row=3, column=1, sticky=W, pady=4)

page2.grid_columnconfigure(3, weight=1)

footer_label2 = Label(page2, text="by Christian", background="white")
footer_label2.grid(column=4, row=4)

mainloop()

